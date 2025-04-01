import openpyxl
import pdfplumber
import time
import re
from multiprocessing import Pool

#=================任务说明
#本脚本用于判断报告是属于A股还是北交所。不同的报告结构不一样


#-----------------定义变量

#定义全局变量
#定义任务
任务报告期目标="2023年报"
任务时间批次="全量"

#定义控制变量
单次任务上限 = 500
开启任务跳过 = True

#-----------------固定变量

#定义pdf文档位置
pdf输入目录 = "/Users/xiami/Desktop/AICode/任务/"+任务报告期目标+"/"+任务时间批次+"/年报源文件/"
任务文档位置 = "/Users/xiami/Desktop/AICode/任务/"+任务报告期目标+"/"+任务时间批次+"/A扫描任务.xlsx"

#定义excel方法
wb = openpyxl.load_workbook(任务文档位置)
ws = wb.active

#-----------------定义函数

def pdf报告类型判断(行号,文件名,pdf输入目录):

    # 定义判断方法
    判断A股方法 = lambda text: (re.search(r"公司简介和主要财务指标", text) is not None)
    判断北交所方法 = lambda text: ((re.search(r"公司概况", text) is not None) and
                                   (re.search(r"财务会计报告", text) is not None))

    # 新增：记录开始时间
    start_time = time.time()

    # 定义文件类型
    结束消息 = "无法判断"
    报告类型 = ""

    #发生错误时跳过
    try:

        #打开pdf文件
        with pdfplumber.open(pdf输入目录+文件名) as pdf:

            # 遍历前60页
            for page in pdf.pages[:60]:

                #提取当前页的文本内容
                提取文案 = page.extract_text()

                #处理提取文案
                提取文案 = 提取文案.replace(" ", "").replace("\n", "")

                #判断是否为A股
                if 判断A股方法(提取文案):
                    报告类型 = "A股"
                    结束消息 = "扫描成功"
                    break
                #判断是否为北交所
                if 判断北交所方法(提取文案):
                    报告类型 = "北交所"
                    结束消息 = "扫描成功"
                    break

    except Exception:
        结束消息 = "文件错误"

    # 计算耗时
    end_time = time.time()
    elapsed_time = end_time - start_time
    print("文件分析完成:" + 文件名 + f"（耗时: {elapsed_time:.2f} 秒）")

    #输出文件类型
    return 行号,结束消息,报告类型

def 判断报告类型_确认任务(开启任务跳过,单次任务上限):
    """
    确定任务范围，前置条件：4，跳过判断：5
    :param 单次任务上限:
    :return: 任务列表
    """

    #定义变量
    任务数量 = 0
    任务列表 = []

    # 遍历每一行（从第二行开始，跳过表头）
    for row in ws.iter_rows(min_row=4, max_col=6):

        #获取关键值
        行号 = row[0].row
        前置条件 = row[4].value
        跳过判断 = row[5].value

        #当开启跳过为True时跳出循环进行下一个循环
        if 开启任务跳过:
            if 跳过判断 is not None:
                continue

        #当前置条件为否是跳出循环
        if 前置条件 != "是":
            continue

        #将任务加入任务列表
        任务列表.append(行号)
        任务数量+=1

        #当单次任务上限超出时跳出循环
        if 任务数量 >= 单次任务上限:
            break

    return 任务列表


#-----------------任务执行


if __name__ == "__main__":

    # # 循环任务(调试)
    # x=0
    # while x <1 :
    #     x+=1

    # 循环任务
    while 判断报告类型_确认任务(开启任务跳过,单次任务上限):

        # 新增：记录开始时间
        单次循环开始时间 = time.time()

        #获取任务列表
        任务列表 = 判断报告类型_确认任务(开启任务跳过,单次任务上限)
        print("任务规划完成，本次循环共计任务数：" + str(len(任务列表)))

        #创建结果表
        扫描结果=[]

        #创建进程池
        进程池=Pool(6)

        # 遍历每一行（从第3行开始，跳过表头）
        for row in ws.iter_rows(min_row=4, max_col=4):

            #判断任务序号是否在任务列表中
            if row[0].row not in 任务列表:
                continue

            #获取关键值
            行号 = row[0].row
            文件名 = row[0].value

            #执行函数
            result = 进程池.apply_async(pdf报告类型判断, (行号,文件名,pdf输入目录))
            扫描结果.append(result)

        进程池.close()
        进程池.join()

        for result in 扫描结果:
            写入序号,结束消息,报告类型 = result.get()

            #判断文件类型
            if 结束消息 == "扫描成功":
                是否下一步="是"
            else:
                是否下一步="否"

            ws.cell(row=写入序号, column=6, value=结束消息)
            ws.cell(row=写入序号, column=7, value=报告类型)
            ws.cell(row=写入序号, column=8, value=是否下一步)
        wb.save(任务文档位置)

        #记录完成时间
        单次循环结束时间 = time.time()
        单次循环耗时 = 单次循环结束时间 - 单次循环开始时间
        print("本次循环任务完成，"f"（耗时: {单次循环耗时:.2f} 秒）")