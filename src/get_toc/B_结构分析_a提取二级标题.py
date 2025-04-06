import pdfplumber
import re
import time
from multiprocessing import Pool
import openpyxl


#=================任务说明
#本脚本用于找到文件的第二节公司简介和主要财务指标，并提取字号。


#-----------------定义变量

#定义全局变量
#定义任务
任务报告期目标="2023年报"
任务时间批次="全量"

#定义控制变量
单次任务上限 = 200
开启任务跳过 = True

#-----------------固定变量

#定义pdf文档位置
pdf输入目录 = "/Users/xiami/Desktop/AICode/任务/"+任务报告期目标+"/"+任务时间批次+"/年报源文件/"
任务文档位置 = "/Users/xiami/Desktop/AICode/任务/"+任务报告期目标+"/"+任务时间批次+"/A扫描任务.xlsx"

#定义excel方法
wb = openpyxl.load_workbook(任务文档位置)
ws = wb.active

#字典



#-----------------定义函数

def 年报扫描_扫描二级标题(行号,pdf名称,pdf输入目录):
    """
        用于公司简介和主要财务指标
        :param 行号:
        :param pdf名称:
        :param pdf输入目录:
        :return: 行号,结束消息,扫描结果
        """

    # 新增：记录开始时间
    start_time = time.time()

    # 定义判断方法
    判断标题方法 = lambda text: (re.search(r"^(一|二|三|四|五|六|七|八|九|十|十一|十二|十三|十四|十五|十六)、\s*(重要会计政策及会计估计|税项|会计政策和会计估计变更以及前期差错更正的说明|(合并)?财务报表(主要)?项目(注释|附注)|研发支出|合并范围的(变更|变动)|在其他主体中的权益)$", text) is not None)

    # 错误跳过
    try:

        # 扫描PDF----------------------------------------
        # 打开pdf
        with (pdfplumber.open(pdf输入目录 + pdf名称) as pdf):


            # 定义初始值避免报错错误
            结束消息="未找到"
            扫描结果 = []

            # 遍历每一页
            for page in pdf.pages:

                # 提取当前页的文本内容
                words = page.extract_words(keep_blank_chars=True, x_tolerance=40)

                # 遍历words取出关键变量
                for word in words:

                    #提取文案
                    word_text = word.get('text')

                    #去掉空格
                    word_text = word_text.replace(" ", "")

                    #判断条件
                    if 判断标题方法(word_text):

                        #整合结果
                        整合结果=[page.page_number,word_text]
                        扫描结果.append(整合结果)

        #结束消息
        结束消息="扫描成功"

    # 错误跳过
    except Exception as e:
        结束消息 = "扫描失败"
        print("扫描失败：" + str(e))

    # 记录完成时间
    end_time = time.time()
    elapsed_time = end_time - start_time
    耗时 = f"（耗时: {elapsed_time:.2f} 秒）"

    # 打印耗时
    print("文件分析完成" + pdf名称 + f"（耗时: {elapsed_time:.2f} 秒）")

    return 行号,结束消息,扫描结果

def 年报扫描_公司简介和主要财务指标_确认任务(开启任务跳过,单次任务上限):
    """
    确定任务范围，前置条件：7，跳过判断：8
    :param 单次任务上限:
    :return: 任务列表
    """

    #定义变量
    任务数量 = 0
    任务列表 = []

    # 遍历每一行（从第四行开始，跳过表头）
    for row in ws.iter_rows(min_row=4, max_col=9):

        #获取关键值
        行号 = row[0].row
        前置条件 = row[7].value
        跳过判断 = row[8].value

        #当开启跳过为True时跳出循环进行下一个循环
        if 开启任务跳过:
            if 跳过判断 is not None:
                continue

        #当前置条件为否是跳出循环
        if 前置条件 != "是":
            continue

        #将任务加入任务列表
        任务列表.append(行号)
        任务数量+=1

        #当单次任务上限超出时跳出循环
        if 任务数量 >= 单次任务上限:
            break

    return 任务列表


#-----------------任务执行


if __name__ == "__main__":

    # # 循环任务(调试)
    # x=0
    # while x <1 :
    #     x+=1

    # 循环任务
    while 年报扫描_公司简介和主要财务指标_确认任务(开启任务跳过,单次任务上限):

        # 新增：记录开始时间
        单次循环开始时间 = time.time()

        #获取任务列表
        任务列表 = 年报扫描_公司简介和主要财务指标_确认任务(开启任务跳过,单次任务上限)
        print("任务规划完成，本次循环共计任务数：" + str(len(任务列表)))

        #创建结果表
        扫描结果=[]

        #创建进程池
        进程池=Pool(6)

        # 遍历每一行（从第3行开始，跳过表头）
        for row in ws.iter_rows(min_row=4, max_col=20):

            #判断任务序号是否在任务列表中
            if row[0].row not in 任务列表:
                continue

            #获取关键值
            行号 = row[0].row
            文件名 = row[0].value
            报告类型 = row[6].value

            #执行函数
            result = 进程池.apply_async(年报扫描_扫描二级标题, (行号,文件名,pdf输入目录))
            扫描结果.append(result)

        进程池.close()
        进程池.join()

        #读取扫描结果
        for result in 扫描结果:
            行号,结束消息,扫描结果 = result.get()

            #判断文件类型
            if 结束消息 == "扫描成功":
                if len(扫描结果) >= 1:
                    是否下一步="是"
                else:
                    是否下一步="否"
            else:
                是否下一步="否"

            #写入结果
            ws.cell(row=行号, column=9, value=结束消息)
            ws.cell(row=行号, column=10, value=str(扫描结果))
            ws.cell(row=行号, column=11, value=是否下一步)

        wb.save(任务文档位置)

        #记录完成时间
        单次循环结束时间 = time.time()
        单次循环耗时 = 单次循环结束时间 - 单次循环开始时间
        print("本次循环任务完成，"f"（耗时: {单次循环耗时:.2f} 秒）")



